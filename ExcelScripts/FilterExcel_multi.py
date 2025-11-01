# FilterExcel_multi.py
# This script filters an Excel file based on a primary email and office match.
# Email and Office key/value pairs are input by the user.
# User can add multiple filters.
# The script writes the original data and filtered data to new Excel files.
# The script styles the Excel files with a dark blue header and zebra striping.
# The script also highlights duplicate rows in the filtered sheets.
# The script also creates a remainder sheet for rows that do not match any filter.

import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
import os
import re

def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize a string to be a valid Excel sheet name:
    - Max 31 characters
    - No: : \ / ? * [ ]
    - No leading/trailing apostrophes
    - Replace disallowed chars with spaces and collapse whitespace
    """
    if name is None:
        name = ""
    # Replace invalid characters with space
    name = re.sub(r'[:\\/?*\[\]]', ' ', str(name))
    # Collapse whitespace
    name = re.sub(r'\s+', ' ', name).strip()
    # Remove leading/trailing apostrophes
    name = name.strip("'")
    # Truncate to 31 chars
    return name[:31] if name else ""

def style_worksheet(ws):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")  # white text
    header_fill = PatternFill("solid", fgColor="156082")  # dark blue background
    stripe1 = PatternFill("solid", fgColor="C0E6F5")  # even rows
    stripe2 = PatternFill("solid", fgColor="FFFFFF")  # odd rows

    # Hide columns A–C
    #for col in ["A", "B", "C"]:
     #   ws.column_dimensions[col].hidden = True

    # Set all column widths to 28
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 28

    # Header formatting
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Zebra striping for data rows (rows 2..max_row)
    for r in range(2, ws.max_row + 1):
        fill = stripe1 if (r % 2 == 0) else stripe2
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def main():
    # -------- FILE SELECTION DIALOG --------
    root = tk.Tk()
    root.withdraw()  # Hide the root window
    file_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )

    if not file_path:
        print("No file selected. Exiting.")
        return

    # -------- CONFIG --------
    email_column = "Primary Email"
    office_column = "Office"

    # -------- READ EXCEL --------
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        messagebox.showerror("Read Error", f"Failed to read Excel file:\n{e}")
        return

    # Normalize columns
    df.columns = df.columns.str.strip()
    lower_cols = {col.lower(): col for col in df.columns}

    # -------- VALIDATE COLUMNS (case-insensitive) --------
    missing_cols = [col for col in [email_column.lower(), office_column.lower()] if col not in lower_cols]
    if missing_cols:
        messagebox.showerror("Missing Columns",
                             f"Error: Missing required column(s): {', '.join(missing_cols)}")
        return

    # Map actual column names
    email_col_actual = lower_cols[email_column.lower()]
    office_col_actual = lower_cols[office_column.lower()]

    # -------- COLLECT MULTIPLE FILTERS --------
    filters = []
    while True:
        email_match = simpledialog.askstring("Primary Email Key", "Enter the email match (required):")
        if email_match is None:  # user cancelled
            # If no filters yet, cancel whole run
            if not filters:
                print("No filters entered. Exiting.")
                return
            else:
                break

        office_match = simpledialog.askstring("Office Key", "Enter the office match (required):")
        if office_match is None:
            if not filters:
                print("No filters entered. Exiting.")
                return
            else:
                break

        email_match = (email_match or "").strip()
        office_match = (office_match or "").strip()

        if not email_match and not office_match:
            messagebox.showwarning("Empty Filter",
                                   "Both Email Key and Office Key are empty. Please enter at least one.")
            continue

        filters.append((email_match, office_match))

        more = messagebox.askyesno("Add Another Filter?",
                                   "Would you like to add another filter?\n\nYes = add another\nNo = continue")
        if not more:
            break

    # -------- PROCESS FILTERS AND WRITE OUTPUT --------
    base, ext = os.path.splitext(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    new_file_path = f"{base}_filtered_{timestamp}{ext}"

    try:
        with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
            # Original sheet
            df.to_excel(writer, sheet_name="Original Data", index=False)

            used_sheet_names = {"Original Data"}
            # For each filter pair, compute filtered DF and write to its own sheet
            for idx, (email_match, office_match) in enumerate(filters, start=1):
                # Build filter (case-insensitive; OR logic)
                mask = pd.Series(False, index=df.index)

                if email_match:
                    mask = mask | df[email_col_actual].astype(str).str.lower().str.contains(email_match.lower(), na=False)
                if office_match:
                    mask = mask | df[office_col_actual].astype(str).str.lower().str.contains(office_match.lower(), na=False)

                filtered_df = df[mask].copy()

                # Determine a user-friendly, valid, and unique sheet name
                base_name = f"Filtered {sanitize_sheet_name(office_match) or sanitize_sheet_name(email_match) or f'Filter {idx}'}"
                sheet_name = base_name
                suffix = 2
                while sheet_name in used_sheet_names or not sheet_name:
                    sheet_name = f"{base_name} ({suffix})"
                    suffix += 1
                used_sheet_names.add(sheet_name)

                filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply styling to all sheets
            wb = writer.book
            for sheet_name in used_sheet_names:
                ws = wb[sheet_name]
                style_worksheet(ws)

        print(f"✅ Original and {len(filters)} filtered sheet(s) written to new file:\n{new_file_path}")
    except Exception as e:
        messagebox.showerror("Write Error", f"Failed to write Excel file:\n{e}")

if __name__ == "__main__":
    main()
