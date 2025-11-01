# FilterExcel_multiBU.py
# This script filters an Excel file based on a primary email and office match.
# Email and Office key/value pairs are hard-coded in the script.
# The script writes the original data and filtered data to new Excel files.
# The script styles the Excel files with a dark blue header and zebra striping.
# The script also highlights duplicate rows in the filtered sheets.
# The script also creates a remainder sheet for rows that do not match any filter.

import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, messagebox
import os
import re

import unicodedata
import string

# --- Duplicate detection config ---
NAME_COLUMNS = ["Name"]  # first existing (case-insensitive) will be used
IGNORE_PUNCTUATION = True
STRIP_DIACRITICS = True

def normalize_name(value: str) -> str:
    """
    Normalize a name for robust duplicate detection:
    - Unicode normalize (NFKC)
    - Lowercase
    - Trim outer whitespace
    - Collapse internal whitespace to single spaces
    - Optionally strip diacritics and punctuation
    """
    if value is None:
        return ""
    # cast to str, normalize unicode
    s = unicodedata.normalize("NFKC", str(value))
    # lowercase
    s = s.lower()
    # strip outer whitespace
    s = s.strip()
    # collapse internal whitespace
    s = " ".join(s.split())

    if STRIP_DIACRITICS:
        # remove combining marks
        s = "".join(
            ch for ch in unicodedata.normalize("NFD", s)
            if unicodedata.category(ch) != "Mn"
        )
        s = unicodedata.normalize("NFKC", s)

    if IGNORE_PUNCTUATION:
        table = str.maketrans("", "", string.punctuation)
        s = s.translate(table)

        # re-collapse just in case removing punctuation left extra spaces
        s = " ".join(s.split())

    return s


# -------- CONFIG: EXTRA "REMAINDER" OUTPUT --------
CREATE_REMAINDER_SHEET = True
REMAINDER_SHEET_NAME = "Remaining (Unmatched)"

# -------- HARD-CODED FILTER DEFINITIONS --------
# Each key becomes an output sheet; each value is a list of (email/office) pairs.
# A row is included in a sheet if it matches ANY pair in that sheet (OR across pairs).
FILTER_DEFINITIONS = {
    # EXAMPLES — edit freely:
    "GBI": [
        {"email": "gil-bar", "office": "105"},
    ],
    "McCoy": [
        {"email": "mccoy", "office": "662"},
        {"email": "mccoy", "office": "818"}
    ],
    "APA": [
        {"email": "apa-conn", "office": "355"},
    ],
    "HCNYE": [
        {"email": "hcnye", "office": "405"},
    ],
    "Airtech": [
        {"email": "airtech", "office": "691"},
    ],
    "GBS": [
        {"email": "gbs", "office": "815"},
    ],
    "Ginns": [
        {"email": "sjginns", "office": "805"},
    ],
    "DMG": [
        {"email": "dmg", "office": "820"},
    ],
    "DynamicFan": [
        {"email": "dynamic", "office": "210"},
    ],
    "JB": [
        {"email": "jbarrow", "office": ""},
    ],
    "NSG ": [
        {"email": "nevada", "office": ""},
    ],
    "APAV": [
        {"email": "apav", "office": "670"},
    ],
    "Ambient": [
        {"email": "ambient-enterprises", "office": ""},
    ]
}

def sanitize_sheet_name(name: str) -> str:
    """
    Sanitize a string to be a valid Excel sheet name:
    - Max 31 characters
    - No: : \ / ? * [ ]
    - No leading/trailing apostrophes
    - Replace disallowed chars with spaces and collapse whitespace
    """
    if name is None:
        name = ""
    name = re.sub(r'[:\\/?*\[\]]', ' ', str(name))
    name = re.sub(r'\s+', ' ', name).strip()
    name = name.strip("'")
    return name[:31] if name else ""

def style_worksheet(ws):
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Font, PatternFill

    header_font = Font(bold=True, color="FFFFFF")          # white text
    header_fill = PatternFill("solid", fgColor="156082")   # dark blue background
    stripe1 = PatternFill("solid", fgColor="C0E6F5")       # even rows
    stripe2 = PatternFill("solid", fgColor="FFFFFF")       # odd rows

    # Hide columns A–C
    #for col in ["A", "B", "C"]:
     #   ws.column_dimensions[col].hidden = True

    # Set all column widths to 28
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 28

    # Header formatting
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Zebra striping for data rows (rows 2..max_row)
    for r in range(2, ws.max_row + 1):
        fill = stripe1 if (r % 2 == 0) else stripe2
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).fill = fill

def main():
    # -------- FILE SELECTION DIALOG --------
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.askopenfilename(
        title="Select Excel file",
        filetypes=[("Excel files", "*.xlsx *.xls")]
    )
    if not file_path:
        print("No file selected. Exiting.")
        return

    # -------- CONFIG (required columns) --------
    email_column = "Email"
    office_column = "BU Code"
    name_column = "Name"  # used for duplicate detection on filter sheets

    # -------- READ EXCEL --------
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        messagebox.showerror("Read Error", f"Failed to read Excel file:\n{e}")
        return

    # Normalize columns and map case-insensitively
    df.columns = df.columns.str.strip()
    lower_cols = {col.lower(): col for col in df.columns}

    # -------- VALIDATE COLUMNS (email + office required, name optional) --------
    required = [email_column.lower(), office_column.lower()]
    missing_cols = [col for col in required if col not in lower_cols]
    if missing_cols:
        messagebox.showerror("Missing Columns",
                             f"Error: Missing required column(s): {', '.join(missing_cols)}")
        return

    email_col_actual = lower_cols[email_column.lower()]
    office_col_actual = lower_cols[office_column.lower()]
    name_col_exists = name_column.lower() in lower_cols
    name_col_actual = lower_cols.get(name_column.lower(), None)

    # -------- PROCESS FILTERS (HARD-CODED) --------
    base, ext = os.path.splitext(file_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    new_file_path = f"{base}_filtered_{timestamp}{ext}"

    try:
        with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
            # Original sheet
            df.to_excel(writer, sheet_name="Original Data", index=False)

            used_sheet_names = {"Original Data"}

            # Track rows matched by ANY filter across ALL sheets
            global_matched_mask = pd.Series(False, index=df.index)

            # Track duplicate rows (by Excel row number) per filter sheet
            duplicates_by_sheet = {}

            # Create filtered sheets per definition
            for raw_sheet_label, pairs in FILTER_DEFINITIONS.items():
                combined_mask = pd.Series(False, index=df.index)

                if not isinstance(pairs, (list, tuple)):
                    pairs = []

                for pair in pairs:
                    email_match = (pair.get("email") or "").strip()
                    office_match = (pair.get("office") or "").strip()

                    # Skip empty pair
                    if not email_match and not office_match:
                        continue

                    pair_mask = pd.Series(False, index=df.index)
                    if email_match:
                        pair_mask = pair_mask | df[email_col_actual].astype(str).str.lower().str.contains(email_match.lower(), na=False)
                    if office_match:
                        pair_mask = pair_mask | df[office_col_actual].astype(str).str.lower().str.contains(office_match.lower(), na=False)

                    combined_mask = combined_mask | pair_mask

                # Update global union
                global_matched_mask = global_matched_mask | combined_mask

                filtered_df = df[combined_mask].copy()

                # Determine safe + unique sheet name
                base_name = sanitize_sheet_name(raw_sheet_label) or "Filtered"
                sheet_name = base_name
                suffix = 2
                while not sheet_name or sheet_name in used_sheet_names:
                    sheet_name = f"{base_name} ({suffix})"
                    suffix += 1
                used_sheet_names.add(sheet_name)

                # Write (even if empty → headers only)
                filtered_df.to_excel(writer, sheet_name=sheet_name, index=False)

                # ----- Duplicate detection by Name (case-insensitive, normalized) -----
                # Only for filter sheets; skip if no name-like column or no rows.
                if not filtered_df.empty:
                    # Find an actual name column in this sheet (case-insensitive)
                    f_lower_cols = {c.lower(): c for c in filtered_df.columns}
                    name_candidate_actual = None
                    for candidate in NAME_COLUMNS:
                        actual = f_lower_cols.get(candidate.lower())
                        if actual:
                            name_candidate_actual = actual
                            break

                    if name_candidate_actual:
                        # Build normalized keys
                        name_series = filtered_df[name_candidate_actual].astype(str)
                        key = name_series.apply(normalize_name)

                        # Flag duplicates (ignore blanks)
                        dup_mask = key.duplicated(keep=False) & (key != "")
                        if dup_mask.any():
                            # Convert to Excel row numbers (data starts at row 2)
                            dup_rows_excel = (dup_mask[dup_mask].index.to_series().astype(int) + 2).tolist()
                            duplicates_by_sheet[sheet_name] = dup_rows_excel


            # -------- REMAINDER SHEET --------
            if CREATE_REMAINDER_SHEET:
                remainder_df = df[~global_matched_mask].copy()

                base_name = sanitize_sheet_name(REMAINDER_SHEET_NAME) or "Remaining"
                sheet_name = base_name
                suffix = 2
                while not sheet_name or sheet_name in used_sheet_names:
                    sheet_name = f"{base_name} ({suffix})"
                    suffix += 1
                used_sheet_names.add(sheet_name)

                remainder_df.to_excel(writer, sheet_name=sheet_name, index=False)
                # Intentionally not doing duplicate highlighting on remainder sheet per request

            # -------- STYLE ALL SHEETS --------
            wb = writer.book
            from openpyxl.styles import PatternFill
            highlight_fill = PatternFill("solid", fgColor="FFFF00")

            for name in used_sheet_names:
                ws = wb[name]
                style_worksheet(ws)

                # After styling, apply duplicate highlighting for filter sheets only
                if name in duplicates_by_sheet:
                    for r in duplicates_by_sheet[name]:
                        # Guard: ensure row exists
                        if r <= ws.max_row:
                            for c in range(1, ws.max_column + 1):
                                ws.cell(row=r, column=c).fill = highlight_fill

        print(
            f"✅ Original, {len(FILTER_DEFINITIONS)} filtered sheet(s)"
            f"{' + remainder' if CREATE_REMAINDER_SHEET else ''} written to:\n{new_file_path}\n"
            f"🔍 Duplicate highlighting: " + (", ".join([f"{k}: {len(v)} row(s)" for k, v in duplicates_by_sheet.items()]) if duplicates_by_sheet else "none")
        )
    except Exception as e:
        messagebox.showerror("Write Error", f"Failed to write Excel file:\n{e}")

if __name__ == "__main__":
    main()
