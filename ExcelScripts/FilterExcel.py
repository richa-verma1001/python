# FilterExcel.py
# This script filters an Excel file based on a primary email and office match.
# Primary Email and Office match are input by the user. 
# The script writes the original data and filtered data to new Excel files.
# The script styles the Excel files with a dark blue header and zebra striping.
# The script hides columns A, B, and C.
# The script sets all column widths to 28.
# The script formats the header row with a white text and dark blue background.
# The script formats the data rows with a zebra striping pattern.


import pandas as pd
from datetime import datetime
import tkinter as tk
from tkinter import filedialog, simpledialog
import os

# -------- FILE SELECTION DIALOG --------
root = tk.Tk()
root.withdraw()  # Hide the root window
file_path = filedialog.askopenfilename(
    title="Select Excel file",
    filetypes=[("Excel files", "*.xlsx *.xls")]
)

if not file_path:
    print("No file selected. Exiting.")
    exit()

# -------- CONFIG --------
email_column = "Primary Email"
office_column = "Office"

# -------- READ EXCEL --------
df = pd.read_excel(file_path)
df.columns = df.columns.str.strip()  # remove spaces

# Convert columns to lowercase for case-insensitive matching
lower_cols = {col.lower(): col for col in df.columns}

# -------- VALIDATE COLUMNS (case-insensitive) --------
missing_cols = [col for col in [email_column.lower(), office_column.lower()] if col not in lower_cols]
if missing_cols:
    print(f"Error: Missing required column(s): {', '.join(missing_cols)}")
    exit()

# Map actual column names
email_col_actual = lower_cols[email_column.lower()]
office_col_actual = lower_cols[office_column.lower()]
email_match = simpledialog.askstring("Primary Email Key", "Enter the email match")
office_match = simpledialog.askstring("Office Key", "Enter the office match")

# -------- FILTER (case-insensitive) --------
filtered_df = df[
    df[email_col_actual].astype(str).str.lower().str.contains(email_match, na=False) |
    df[office_col_actual].astype(str).str.lower().str.contains(office_match, na=False)
]

# -------- WRITE TO NEW FILE --------
base, ext = os.path.splitext(file_path)
timestamp = datetime.now().strftime("%Y%m%d_%H%M")
new_file_path = f"{base}_filtered_{timestamp}{ext}"


from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

with pd.ExcelWriter(new_file_path, engine="openpyxl") as writer:
    # Write both sheets
    df.to_excel(writer, sheet_name="Original Data", index=False)
    filtered_df.to_excel(writer, sheet_name="Filtered " + office_match, index=False)

    # Access workbook and worksheets
    wb = writer.book
    ws1 = writer.sheets["Original Data"]
    ws2 = writer.sheets["Filtered " + office_match]

    # Define header style
    header_font = Font(bold=True, color="FFFFFF")  # white text
    header_fill = PatternFill("solid", fgColor="156082")  # dark blue background
    stripe1 = PatternFill("solid", fgColor="C0E6F5")         # even data rows
    stripe2 = PatternFill("solid", fgColor="FFFFFF")         # odd data rows

    # Apply header style to both sheets
    for ws in [ws1, ws2]:
        # Hide columns A–C
        for col in ["A", "B", "C"]:
            ws.column_dimensions[col].hidden = True

        # Set all column widths to 28 
    
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 28

        # Format header row (row 1)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        # Zebra striping for data rows (rows 2 .. max_row)
        # Even-numbered rows -> #C0E6F5, odd-numbered rows -> #FFFFFF
        for r in range(2, ws.max_row + 1):
            fill = stripe1 if (r % 2 == 0) else stripe2
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).fill = fill

print(f"✅ Original and filtered data written to new file: {new_file_path}")

